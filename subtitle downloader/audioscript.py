#!/usr/bin/env python3
from youtube_transcript_api import YouTubeTranscriptApi
import youtube_transcript_api
import urllib.parse as p
import argparse
import sys
from openpyxl import Workbook
import random
import string


def get_video_id_by_url(url):
    """
    Return the Video ID from the video `url`
    """
    # split URL parts
    parsed_url = p.urlparse(url)
    # get the video ID by parsing the query of the URL
    video_id = p.parse_qs(parsed_url.query).get("v")
    if video_id:
        return video_id[0]
    else:
        # raise Exception(f"Wasn't able to parse video URL: {url}")
        print("[INVALID URL]")
        print(". Make sure to provide a valid youtube video url")
        sys.exit(5)


def sanitize_filename(filename):
    # Define a set of characters that are not allowed in filenames
    illegal_chars = '<>:"/\\|?*'

    # Replace illegal characters with an empty string
    sanitized_filename = "".join(c if c not in illegal_chars else "" for c in filename)

    return sanitized_filename


def positive_int(value):
    """
    Handler function used by `argparser` to only allow positive
     integers as arguments.
    """
    ivalue = int(value)
    if ivalue < 0:
        raise argparse.ArgumentTypeError(f"{value} is not a valid positive integer")
    return ivalue


#### Argument parser to get cli arguments
parser = argparse.ArgumentParser(
    prog="audioscript.py",
    description="Youtube videos audioscript(transcript) parser. It receives a valid youtube video url, downloads transcript, parses it and stores it in form of an spreadsheet with filename 'audioscript-{3_random_letter}.xlsx'.",
)
parser.add_argument(
    "url", metavar="URL", help="URL of youtube video to retrieve the audioscript"
)
parser.add_argument(
    "-n",
    default=0,
    type=positive_int,
    help="Number of words per line in the resulting parsed spreadsheet. Use '0' to use original sentences and timeline. (Defaults to '0' if not specified)",
)

args = parser.parse_args()
video_id = get_video_id_by_url(args.url)
n = args.n
# assigning srt variable with the list
# of dictionaries obtained by the get_transcript() function


######## Youtube api call to fetch video title to use in spreadsheet filename
try:
    from googleapiclient.discovery import build

    api_key = "AIzaSyAATZR1wuL-JUj0tQivp2P9p8CRu8aUNgs"
    youtube = build("youtube", "v3", developerKey=api_key)

    video_response = youtube.videos().list(part="snippet", id=video_id).execute()

    items = video_response.get("items")[0]
    # get the snippet, statistics & content details from the video response
    snippet = items["snippet"]
    video_title = snippet["title"]

    filename = f"{video_title}-audioscript"
except Exception:
    filename = f"audioscript-{sanitize_filename(video_id)}"
    # "".join(randfom.SystemRandom().choices(string.ascii_lowercase, k=3))
else:
    print(f"[+] Video Title: {video_title}")

### Retrieve available audioscripts/transcripts (performs toplevel error handling)
try:
    print(f"[Requested Audioscript for the video id `{video_id}`]")
    transcript_list = YouTubeTranscriptApi.list_transcripts(video_id)
except youtube_transcript_api._errors.TranscriptsDisabled:
    print("[SUBTITLES ARE DISABLED FOR THIS VIDEO]")
    sys.exit(2)
except Exception:
    print(
        """[-] An ERROR Happend, possible couses:
    1. VPN Access - In IRAN vpn access is required to access youtube servers
    2. Connection Issue - Check you're Internet connection
    3. Dependency Issue - Python's `requests` and `youtube_transcript_api`
     modules are required
        """
    )
    sys.exit(4)

""" `YouTubeTranscriptApi.list_transcripts` object holds
 list of different types of audioscripts in these three
 objects:
transcript_list._manually_created_transcripts
transcript_list._generated_transcripts
transcript_list._translation_languages"""

##### Get language codename from user and retrieve audioscript (handles differents types of audioscript accordingly)
print(transcript_list)
lang = input(" [SELECT LANGUAGE USING CODENAME]\n> ")
if lang in transcript_list._manually_created_transcripts:
    print("[RETRIEVING MANUAL GENERATED AUDIOSCRIPT]")
    transcript = transcript_list.find_transcript([lang]).fetch()
elif lang in transcript_list._generated_transcripts:
    print("[RETRIEVING AUTO GENERATED AUDIOSCRIPT]")
    transcript = transcript_list.find_transcript([lang]).fetch()
elif lang in [
    language["language_code"] for language in transcript_list._translation_languages
]:
    default_lang = next(iter(transcript_list))
    print("[RETRIEVING TRANSLATED AUDIOSCRIPT]")
    transcript = (
        transcript_list.find_transcript(["en", default_lang]).translate(lang).fetch()
    )
else:
    print("[INCORRECT OR UNAVAILABLE LANGUAGE CODE]")
    sys.exit(3)

############ Finish constructing filename by adding language name to it
lang_name = [
    language["language"]
    for language in transcript_list._translation_languages
    if language["language_code"] == lang
][0]
filename = f"{filename}_{lang_name}.xlsx"
filename = sanitize_filename(filename)

#### Define spreadsheet object and format in order to save the data
workbook = Workbook()
sheet = workbook.active
if n == 0:
    sheet.cell(row=1, column=1).value = "start (seconds)"
    sheet.cell(row=1, column=2).value = "duration (seconds)"
    sheet.cell(row=1, column=3).value = "text"
    for index, sentence in enumerate(transcript):
        sheet.cell(row=index + 2, column=1).value = sentence["start"]
        sheet.cell(row=index + 2, column=2).value = sentence["duration"]
        sheet.cell(row=index + 2, column=3).value = sentence["text"]
else:
    sheet.cell(row=1, column=1).value = f"{n} word per line (n={n})"
    word_list = (
        " ".join([sentence["text"] for sentence in transcript])
        .replace("\n", " ")
        .replace(",", " ")
        .replace(".", " ")
        .replace('"', "")
        .replace("'", "")
        .replace("  ", " ")
        .replace("  ", " ")
        .split(" ")
    )
    n_list = [" ".join(word_list[i : i + n]) for i in range(0, len(word_list), n)]
    for index, line in enumerate(n_list):
        sheet.cell(row=index + 2, column=1).value = line


try:
    workbook.save(filename=filename)
except Exception:
    print("[FAILED TO SAVE THE SPREADSHEET]")
else:
    print(f"[+] Saved Spreadsheet to '{filename}'")
