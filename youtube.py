#!/usr/bin/env python3
from googleapiclient.discovery import build
import urllib.parse as p
import re
from openpyxl import Workbook
import sys

"""
Output in Two spreadsheets:
------ 1. First
URL
Title
Duration/Publish Time
Like Count
Comment Count
View Count
Description
------ 2. Second
Channel Video Count
Subscribe Count
Channel Time
"""
videos_dict = dict()
channel_dict = dict()


def youtube_authenticate():
    """
    Authenticates to Youtube API with a valid token. This is required to
     communicate to the api. (A valid token can be obtained using a valid
      gmail account. Refer to official google docs.)
    """
    api_key = "AIzaSyAATZR1wuL-JUj0tQivp2P9p8CRu8aUNgs"
    return build("youtube", "v3", developerKey=api_key)


def get_video_id_by_url(url):
    """
    Return the Video ID from the video `url`
    """
    # split URL parts
    parsed_url = p.urlparse(url)
    # get the video ID by parsing the query of the URL
    video_id = p.parse_qs(parsed_url.query).get("v")
    if video_id:
        return video_id[0]
    else:
        raise Exception(f"Wasn't able to parse video URL: {url}")


def get_video_details(youtube, **kwargs):
    """
    Performs an api call to google server and request 3 portions of information
    each corresponding a subset of attributes associated with a youtube channel.
    (It returns dict-like object)
    """
    return (
        youtube.videos()
        .list(part="snippet,contentDetails,statistics", **kwargs)
        .execute()
    )


#
def retrieve_video_infos(video_response):
    """
    This function takes dict-like object returned by `get_video_details` which
    contains indivisual data attributes associated with a youtube video and
     extracts and stores desired data items in a global dictionary.
    """
    items = video_response.get("items")[0]
    # get the snippet, statistics & content details from the video response
    snippet = items.get("snippet", "NA")
    statistics = items.get("statistics", "NA")
    content_details = items.get("contentDetails", "NA")
    # get infos from the snippet
    # channel_title = snippet["channelTitle"]
    title = snippet.get("title", "NA")
    description = snippet.get("description", "NA")
    publish_time = snippet.get("publishedAt", "NA")
    # get stats infos
    comment_count = statistics.get("commentCount", "NA")
    like_count = statistics.get("likeCount", "NA")
    view_count = statistics.get("viewCount", "NA")
    # get duration from content details
    duration = content_details.get("duration", "NA")
    duration_str = duration
    videos_dict.update(
        {
            items["id"]: {
                "title": title,
                "duration": duration,
                "publish_time": publish_time,
                "view_count": view_count,
                "like_count": like_count,
                "comment_count": comment_count,
                "description": description,
            }
        }
    )


def search(youtube, **kwargs):
    """
    This api call is used to find channel id by its name. This func
    is called by `get_channel_id_by_url` func in the process of obtaining
    channel id by many forms of its url.
    """
    return youtube.search().list(part="snippet", **kwargs).execute()


def parse_channel_url(url):
    """
    This function takes channel `url` to check whether it includes a
    channel ID, user ID or channel name, is used by `get_channel_id_by_url` func.
    """
    path = p.urlparse(url).path
    # id = path.split("/")[-1]
    if "/c/" in path:
        id = path.split("/c/")[1].split("/")[0]
        return "c", id
    elif "/channel/" in path:
        id = path.split("/channel/")[1].split("/")[0]
        return "channel", id
    elif "/user/" in path:
        id = path.split("/user/")[1].split("/")[0]
        return "user", id
    elif "/@" in path:
        id = path.split("@")[1].split("/")[0]
        return "c", id
    else:
        print("[INVALID URL CHANNEL]")
        sys.exit(2)


def get_channel_id_by_url(youtube, url):
    """
    Returns channel ID of a given `id` and `method`
    - `method` (str): can be 'c', 'channel', 'user'
    - `id` (str): if method is 'c', then `id` is display name
        if method is 'channel', then it's channel id
        if method is 'user', then it's username
    """
    # parse the channel URL
    method, id = parse_channel_url(url)
    if method == "channel":
        # if it's a channel ID, then just return it
        return id
    elif method == "user":
        # if it's a user ID, make a request to get the channel ID
        response = get_channel_details(youtube, forUsername=id)
        items = response.get("items")
        if items:
            channel_id = items[0].get("id")
            return channel_id
    elif method == "c":
        # if it's a channel name, search for the channel using the name
        # may be inaccurate
        response = search(youtube, q=id, maxResults=1)
        items = response.get("items")
        if items:
            channel_id = items[0]["snippet"]["channelId"]
            return channel_id
    raise Exception(f"Cannot find ID:{id} with {method} method")


def get_channel_videos(youtube, **kwargs):
    """
    This api call returns a dict-like object containing an specific
     subset of youtube video object. This func is called by 'retrieve_channel_infos'
     func many times to retrieve all channel videos'.
    """
    return youtube.search().list(**kwargs).execute()


def get_channel_details(youtube, **kwargs):
    """
    Using the channel id, requests for 3 portions of info about channel by
    performing an api call. (returns a dict-like object containing data attrs
    associated with the channel)
    """
    return (
        youtube.channels()
        .list(part="statistics,snippet,contentDetails", **kwargs)
        .execute()
    )


def retrieve_channel_infos(response):
    """
    Extract channel infos from dict-like object returned by `get_channel_details`
    func. Then recursively requests channel videos', and iterates through the list
    to fetch indivisual videos' info.
    """

    snippet = response["items"][0]["snippet"]
    statistics = response["items"][0]["statistics"]
    channel_country = snippet.get("country", "NA")

    channel_description = snippet.get("description", "NA")
    channel_creation_date = snippet.get("publishedAt", "NA")
    channel_title = snippet.get("title", "NA")
    channel_subscriber_count = statistics.get("subscriberCount", "NA")
    channel_video_count = statistics.get("videoCount", "NA")
    channel_view_count = statistics.get("viewCount", "NA")
    channel_dict.update(
        {
            "id": channel_id,
            "title": channel_title,
            "country": channel_country,
            "creation_date": channel_creation_date,
            "subscriber_count": channel_subscriber_count,
            "view_count": channel_view_count,
            "description": channel_description,
        }
    )
    # the following is grabbing channel videos
    # number of pages you want to get
    n_pages = 20
    # counting number of videos grabbed
    n_videos = 0
    next_page_token = None
    next_page_flag = True
    print("[+] Processing channel Videos ")
    while next_page_flag:
        params = {
            "part": "snippet",
            "q": "",
            "channelId": channel_id,
            "type": "video",
        }
        if next_page_token:
            params["pageToken"] = next_page_token
        res = get_channel_videos(youtube, **params)
        channel_videos = res.get("items")
        for video in channel_videos:
            print(".", end="", flush=True)
            n_videos += 1
            video_id = video["id"]["videoId"]
            # easily construct video URL by its ID
            video_response = get_video_details(youtube, id=video_id)
            # print the video details
            retrieve_video_infos(video_response)
        # if there is a next page, then add it to our parameters
        # to proceed to the next page
        if "nextPageToken" in res:
            next_page_token = res["nextPageToken"]
        else:
            next_page_flag = False
            print()  # Just to go to next line


def store_spreadsheet():
    """
    All the code to actually format and save to disk desired output
     spreadsheets is here.
     (Global `video_dict` and `channel_dict` are used here.)
    """
    workbook = Workbook()
    sheet = workbook.active

    for index, column_tag in enumerate(channel_dict):
        sheet.cell(row=1, column=index + 1).value = column_tag
    for index, column_value in enumerate(channel_dict.values()):
        sheet.cell(row=2, column=index + 1).value = column_value

    filename = "{}_channel-info.xlsx".format(channel_dict["title"].lower())
    try:
        workbook.save(filename=filename)
    except Exception:
        print("[FAILED TO SAVE THE SPREADSHEET]")
    else:
        print(f"[+] Saved Spreadsheet to '{filename}'")

    workbook.remove(sheet)
    sheet = workbook.create_sheet()
    sheet.cell(row=1, column=1).value = "url"
    for index, column_tag in enumerate(next(iter(videos_dict.values()))):
        sheet.cell(row=1, column=index + 2).value = column_tag
    for video_index, video in enumerate(videos_dict):
        sheet.cell(
            row=video_index + 2, column=1
        ).value = f"https://www.youtube.com/watch?v={video}"
        for column_index, column_value in enumerate(videos_dict[video].values()):
            sheet.cell(
                row=video_index + 2, column=column_index + 2
            ).value = column_value
    filename = "{}_videos-info.xlsx".format(channel_dict["title"].lower())
    try:
        workbook.save(filename=filename)
    except Exception:
        print("[FAILED TO SAVE THE SPREADSHEET]")
    else:
        print(f"[+] Saved Spreadsheet to '{filename}'")


if __name__ == "__main__":
    # authenticate to YouTube API
    print("[AUTHENTICATING with YOUTUBE API...]")
    try:
        youtube = youtube_authenticate()
    except Exception:
        print(
            """[-] An ERROR Happend, possible couses:
    1. VPN Access - In IRAN vpn access is required to access youtube servers
    2. Connection Issue - Check you're Internet connection
    3. Dependency Issue - Python's `googleapiclient` module is required
        """
        )
        sys.exit(4)

    # parse video ID from URL
    # video_id = get_video_id_by_url(video_url)
    # make API call to get video info
    # response = get_video_details(youtube, id=video_id)
    # print extracted video infos
    # print_video_infos(response)

    channel_url = input(" [+] Please enter the Youtube channel URL:\n > ")

    # get the channel ID from the URL
    channel_id = get_channel_id_by_url(youtube, channel_url)
    # get the channel details
    response = get_channel_details(youtube, id=channel_id)
    # print extracted video infos
    print("[RETRIVING CHANNEL INFO...]")
    retrieve_channel_infos(response)

    try:
        store_spreadsheet()
    except Exception:
        print("[FAILED TO GENERATE THE SPREADSHEETS]")
        sys.exit(6)
